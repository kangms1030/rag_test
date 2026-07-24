"""엑셀 모범 질답 데이터를 data/faq.json 으로 정규화한다.

원본 엑셀은 읽기만 하며 수정하지 않는다. 엑셀 변경 시 재실행:

    conda activate intern_chatbot
    cd C:\\Users\\minsoo\\Desktop\\아이티지엔 인턴\\챗봇
    python -m chatbot_demo.scripts.import_excel

원본이 다른 경로면 --excel 로 지정. 출력은 --out.

엑셀 quirk 처리:
  - 헤더 변형: 스쿨넷 "질문유형"(공백 없음) vs 타 시트 "질문 유형" → 공백 제거해 매칭
  - 스쿨넷 G열(헤더 없음, '?' 등) → 무시(정의된 헤더 컬럼만 사용)
  - 유무선통합관제 유령 빈 행(52~103) → 질문/답변 빈 행 skip
  - 질문 유형 빈 셀 → None 유지
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import load_workbook

# 패키지 임포트가 가능하도록(직접 실행 대비)
_PKG_ROOT = Path(__file__).resolve().parents[1]
if str(_PKG_ROOT.parent) not in sys.path:
    sys.path.insert(0, str(_PKG_ROOT.parent))

from chatbot_demo.scenario.matcher import normalize_text  # noqa: E402

DEFAULT_EXCEL = (
    _PKG_ROOT.parent
    / "쓰잘데기"
    / "★(클로드) 장애 상담 데이터_합본_정리-질문유형추가_답변정리_0710.xlsx"
)
DEFAULT_OUT = _PKG_ROOT / "data" / "faq.json"

EXPECTED_SHEETS = ["스쿨넷", "학내망", "무선망", "유무선통합관제"]

# 근거 파일명 셀을 여러 파일로 분리하는 구분자
_SPLIT_CHARS = ["\n", ";", ",", "/"]


def _norm_header(h) -> str:
    """헤더 정규화: 공백 제거 후 비교용."""
    if h is None:
        return ""
    return str(h).replace(" ", "").strip()


def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _split_source_files(raw) -> list[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    parts = [text]
    for ch in _SPLIT_CHARS:
        nxt: list[str] = []
        for p in parts:
            nxt.extend(p.split(ch))
        parts = nxt
    seen: list[str] = []
    for p in parts:
        p = p.strip()
        if p and p not in seen:
            seen.append(p)
    return seen


def _column_map(header_row) -> dict[str, int]:
    """정규화 헤더명 → 0-based 컬럼 인덱스."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _norm_header(cell)
        if key and key not in mapping:
            mapping[key] = idx
    return mapping


def import_excel(excel_path: Path, out_path: Path) -> dict:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    entries: list[dict] = []
    per_sheet: dict[str, int] = {}
    skipped_empty = 0
    seen_norm: dict[str, str] = {}  # norm_q -> first id (중복 감지)
    dup_warnings: list[str] = []

    for sheet_name in EXPECTED_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"[import_excel] 시트를 찾을 수 없음: {sheet_name}")
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            per_sheet[sheet_name] = 0
            continue
        cmap = _column_map(header)

        # 컬럼 인덱스(헤더 변형 흡수)
        c_no = cmap.get("No")
        c_qtype = cmap.get("질문유형")            # "질문유형"/"질문 유형" 모두 정규화됨
        c_ftype = cmap.get("장애유형")            # "장애 유형"
        c_q = cmap.get("질문")
        c_a = cmap.get("답변")
        c_src = cmap.get("질문답변생성근거파일명")

        if c_q is None or c_a is None:
            raise SystemExit(
                f"[import_excel] {sheet_name}: 질문/답변 컬럼을 찾지 못함. 헤더={header}"
            )

        count = 0
        for r_idx, row in enumerate(rows, start=2):  # 헤더가 1행, 데이터 2행부터
            question = _clean(row[c_q]) if c_q < len(row) else None
            answer = _clean(row[c_a]) if c_a < len(row) else None
            if not question or not answer:
                skipped_empty += 1
                continue

            no_val = None
            if c_no is not None and c_no < len(row):
                raw_no = row[c_no]
                try:
                    no_val = int(raw_no) if raw_no is not None else None
                except (ValueError, TypeError):
                    no_val = None

            qtype = _clean(row[c_qtype]) if c_qtype is not None and c_qtype < len(row) else None
            ftype = _clean(row[c_ftype]) if c_ftype is not None and c_ftype < len(row) else None
            src = row[c_src] if c_src is not None and c_src < len(row) else None

            norm_q = normalize_text(question)
            entry_id = f"{sheet_name}:{r_idx}"
            if norm_q in seen_norm:
                dup_warnings.append(
                    f"중복 정규화 질문: {entry_id} == {seen_norm[norm_q]} ({norm_q[:30]})"
                )
            else:
                seen_norm[norm_q] = entry_id

            entries.append(
                {
                    "id": entry_id,
                    "sheet": sheet_name,
                    "row": r_idx,
                    "no": no_val,
                    "question_type": qtype,
                    "fault_type": ftype,
                    "question": question,
                    "question_normalized": norm_q,
                    "answer": answer,
                    "source_files": _split_source_files(src),
                }
            )
            count += 1
        per_sheet[sheet_name] = count

    wb.close()

    kst = timezone(timedelta(hours=9))
    payload = {
        "version": 1,
        "generated_at": datetime.now(kst).isoformat(),
        "source_file": excel_path.name,  # 파일명만(절대경로 노출 금지)
        "entry_count": len(entries),
        "per_sheet": per_sheet,
        "entries": entries,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"[import_excel] 총 {len(entries)}건 → {out_path}")
    for name in EXPECTED_SHEETS:
        print(f"   - {name}: {per_sheet.get(name, 0)}건")
    print(f"[import_excel] 빈 질문/답변 skip: {skipped_empty}건")
    if dup_warnings:
        print(f"[import_excel] 경고: 중복 정규화 질문 {len(dup_warnings)}건")
        for w in dup_warnings:
            print("     " + w)
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="엑셀 모범 질답 → faq.json")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    if not args.excel.exists():
        raise SystemExit(f"[import_excel] 엑셀 파일 없음: {args.excel}")
    import_excel(args.excel, args.out)


if __name__ == "__main__":
    main()
