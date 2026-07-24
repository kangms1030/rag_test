"""엑셀 정규화 결과 및 재생성 스크립트 검증."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from chatbot_demo.scripts.import_excel import import_excel


def test_faq_json_loads_and_has_expected_counts():
    faq_path = Path(__file__).resolve().parents[1] / "data" / "faq.json"
    data = json.loads(faq_path.read_text(encoding="utf-8"))
    assert data["entry_count"] == 236
    assert data["per_sheet"] == {
        "스쿨넷": 52, "학내망": 50, "무선망": 84, "유무선통합관제": 50,
    }
    # 절대경로 노출 금지 — source_file 은 파일명만
    assert "\\" not in data["source_file"] and "/" not in data["source_file"]
    for e in data["entries"]:
        assert e["question"] and e["answer"]
        assert e["id"] == f"{e['sheet']}:{e['row']}"


def test_import_handles_header_variants_and_ghost_rows(tmp_path):
    """스쿨넷 '질문유형'(공백없음)/타 시트 '질문 유형', 유령 빈 행 처리."""
    wb = Workbook()
    # 스쿨넷: 헤더 '질문유형' + G열 쓰레기
    ws1 = wb.active
    ws1.title = "스쿨넷"
    ws1.append(["No", "질문유형", "장애 유형", "질문", "답변", "질문 답변 생성 근거 파일명", None])
    ws1.append([1, "일반질문", "개념", "스쿨넷이 뭐예요?", "스쿨넷은 ...", "a.pdf 1쪽", "?"])
    # 학내망: 헤더 '질문 유형'(공백)
    ws2 = wb.create_sheet("학내망")
    ws2.append(["No", "질문 유형", "장애 유형", "질문", "답변", "질문 답변 생성 근거 파일명"])
    ws2.append([1, "장애", "연결", "인터넷이 안 돼요", "순서대로...", "b.pdf, c.pdf"])
    ws2.append([2, None, "무선망", "와이파이 비번?", "비번은...", "d.pdf"])  # 빈 질문유형
    ws3 = wb.create_sheet("무선망")
    ws3.append(["No", "질문 유형", "장애 유형", "질문", "답변", "질문 답변 생성 근거 파일명"])
    ws4 = wb.create_sheet("유무선통합관제")
    ws4.append(["No", "질문 유형", "장애 유형", "질문", "답변", "질문 답변 생성 근거 파일명"])
    ws4.append([1, "일반질문", "로그인", "로그인?", "weiss...", "e.pdf"])
    ws4.append([None, None, None, None, None, None])  # 유령 빈 행

    xlsx = tmp_path / "sample.xlsx"
    wb.save(xlsx)
    out = tmp_path / "faq.json"
    payload = import_excel(xlsx, out)

    assert payload["entry_count"] == 4  # 유령 빈 행 skip
    by_id = {e["id"]: e for e in payload["entries"]}
    # 헤더 변형 흡수: 스쿨넷 질문유형 파싱됨
    assert by_id["스쿨넷:2"]["question_type"] == "일반질문"
    # 빈 질문유형 → None
    assert by_id["학내망:3"]["question_type"] is None
    # 근거 다중 파일 분리
    assert by_id["학내망:2"]["source_files"] == ["b.pdf", "c.pdf"]
