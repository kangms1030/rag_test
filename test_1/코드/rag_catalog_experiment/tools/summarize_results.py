"""outputs/ 의 원본 JSON에서 사람이 읽을 결과 뷰를 생성한다.

    python -m tools.summarize_results

읽는 곳: outputs/comparisons/compare_*.json, outputs/evaluations/thresholds_*.json
쓰는 곳: results/00_SUMMARY.md, results/01_질문별_답변비교.md, results/results.xlsx

원본은 수정하지 않는다. 실험을 다시 돌린 뒤 이 스크립트를 재실행하면 results/ 가 갱신된다.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = ROOT / "outputs"
RESULTS = ROOT / "results"
EVAL_SETS = ROOT / "eval_sets"

MODE_LABEL = {
    "catalog": "카탈로그 사용",
    "no_catalog": "카탈로그 미사용",
    "filename_only": "파일명만",
}
MODE_ORDER = ["catalog", "no_catalog", "filename_only"]

# 실험(depth)별로 어떤 평가셋을 썼는지. 기대 답변을 붙이는 데 쓴다.
EVAL_SET_BY_SOURCE = {
    "sample_qa": "qa_sample_20_stratified.json",
    "human": "human_20.json",
}


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def pick_thresholds() -> dict[str, dict]:
    """모드 -> {stat, gate_attr, source}. 모드별로 유효한 최신 측정을 고른다.

    thresholds 실행은 모드 단위로 독립이고, 부분 재실행 결과가 남아 있다.
    예컨대 초기 실행은 filename_only 의 점수를 하나도 수집하지 못해
    (n_relevant=0) 통계가 전부 null이고, 이후 그 모드만 다시 돌린 파일이 있다.
    그래서 파일 하나를 통째로 고르지 않고, 모드마다 n_relevant>0 인 가장 최근
    측정을 쓴다. 어느 파일에서 왔는지는 출처로 함께 남긴다.
    """
    picked: dict[str, dict] = {}
    # 이름순 = 시간순. 최신부터 훑어 첫 유효 측정을 채택한다.
    for path in sorted((OUTPUTS / "evaluations").glob("thresholds_*.json"), reverse=True):
        data = load_json(path)
        for mode, stat in data.get("by_mode", {}).items():
            if mode in picked:
                continue
            if not stat.get("n_relevant"):
                continue  # 점수를 못 모은 실패 측정
            picked[mode] = {
                "stat": stat,
                "gate_attr": data["gate_attr"][mode],
                "source": path.relative_to(ROOT).as_posix(),
            }
    return picked


def fmt(value: Any, digits: int = 3) -> str:
    if value is None:
        return "-"
    if isinstance(value, float):
        return f"{value:.{digits}f}"
    return str(value)


def load_expected_answers() -> dict[str, dict]:
    """평가셋의 id -> 항목. 기대 답변/기대 문서를 답변 비교에 곁들이기 위함."""
    expected: dict[str, dict] = {}
    for filename in EVAL_SET_BY_SOURCE.values():
        path = EVAL_SETS / filename
        if not path.exists():
            continue
        for item in load_json(path):
            expected[item["id"]] = item
    return expected


def collect_compares() -> dict[str, dict]:
    """depth -> 가장 최근 compare 결과."""
    by_depth: dict[str, dict] = {}
    for path in sorted((OUTPUTS / "comparisons").glob("compare_*.json")):
        data = load_json(path)
        data["_path"] = path
        by_depth[data["depth"]] = data  # 이름순이라 뒤에 오는 게 최신
    return by_depth


def per_item_rows(compare: dict) -> list[dict]:
    """compare JSON의 mode별 per_item을 (질문 × 모드) 평탄한 행으로 편다."""
    rows: list[dict] = []
    for mode in MODE_ORDER:
        result = compare["full_results"].get(mode)
        if not result:
            continue
        for item in result["per_item"]:
            rows.append(item)
    return rows


def group_by_question(compare: dict) -> dict[str, dict[str, dict]]:
    """질문 id -> {mode: per_item}. 3모드 나란히 보기용."""
    grouped: dict[str, dict[str, dict]] = {}
    for item in per_item_rows(compare):
        grouped.setdefault(item["id"], {})[item["mode"]] = item
    return grouped


# ---------------------------------------------------------------- 00_SUMMARY

ANSWER_HEADERS = [
    "모드",
    "답변률",
    "정확도 (token_f1)",
    "키워드 재현율",
    "평균 소요시간(초)",
    "VLM 페이지(비용)",
    "평균 선정 문서 수",
]
DOCS_HEADERS = [
    "모드",
    "문서 재현율 doc_recall@k",
    "문서 MRR",
    "무관 질문 거절 정확도",
    "평균 소요시간(초)",
    "평균 선정 문서 수",
]


def answer_summary_rows(compare: dict) -> list[list[str]]:
    rows = []
    for mode in MODE_ORDER:
        s = compare["summary"][mode]["all"]
        rows.append(
            [
                f"{MODE_LABEL[mode]} (`{mode}`)",
                f"{s['answered_count']}/{s['total_questions']}",
                fmt(s["answer_token_f1_avg"]),
                fmt(s["expected_answer_keyword_recall_avg"]),
                fmt(s["avg_elapsed_seconds"], 1),
                fmt(s["avg_vlm_pages_required"], 1),
                fmt(s["avg_selected_doc_count"], 2),
            ]
        )
    return rows


def mean_doc_mrr(compare: dict, mode: str) -> float | None:
    """구버전 compare JSON은 집계에 doc_mrr가 없다. per_item에서 직접 평균낸다."""
    aggregate = compare["summary"][mode]["all"]
    if aggregate.get("doc_mrr") is not None:
        return aggregate["doc_mrr"]
    values = [
        item["doc_mrr"]
        for item in compare["full_results"][mode]["per_item"]
        if item.get("doc_mrr") is not None
    ]
    return sum(values) / len(values) if values else None


def docs_summary_rows(compare: dict) -> list[list[str]]:
    rows = []
    for mode in MODE_ORDER:
        s = compare["summary"][mode]["all"]
        rows.append(
            [
                f"{MODE_LABEL[mode]} (`{mode}`)",
                fmt(s["doc_recall_at_k"]),
                fmt(mean_doc_mrr(compare, mode)),
                fmt(s["rejection_accuracy"]),
                fmt(s["avg_elapsed_seconds"], 1),
                fmt(s["avg_selected_doc_count"], 2),
            ]
        )
    return rows


def md_table(headers: list[str], rows: list[list[str]]) -> str:
    out = ["| " + " | ".join(headers) + " |"]
    out.append("|" + "|".join(["---"] * len(headers)) + "|")
    for row in rows:
        out.append("| " + " | ".join(row) + " |")
    return "\n".join(out)


def write_summary(compares: dict[str, dict], thresholds: dict[str, dict]) -> None:
    answer = compares.get("answer")
    docs = compares.get("docs")
    lines: list[str] = ["# 실험 결과 요약", ""]

    if answer:
        n = answer["eval_item_count"]
        catalog_answer = answer["summary"]["catalog"]["all"]
        no_catalog_answer = answer["summary"]["no_catalog"]["all"]
        lines += [
            "## 한 줄 결론",
            "",
            f"카탈로그를 쓰면 답변이 **{fmt(no_catalog_answer['avg_elapsed_seconds'] / catalog_answer['avg_elapsed_seconds'], 1)}배 빠르고 "
            f"정확도(token_f1)는 {fmt(catalog_answer['answer_token_f1_avg'] / no_catalog_answer['answer_token_f1_avg'], 1)}배 높지만, "
            f"{n}문항 중 {catalog_answer['answered_count']}문항에만 답한다.** "
            "카탈로그를 끄면 거의 모든 질문에 답하는 대신 답변 품질이 떨어지고 오래 걸린다.",
            "",
            "### ⚠️ 결론 내기 전에 반드시 읽을 것",
            "",
            "**세 실험은 같은 질문셋을 쓰지 않는다.** 실험 A는 `qa_*` 질문, 실험 B·C는 `human_*` 질문이다. "
            "그래서 표를 가로질러 비교하면 안 된다.",
            "",
            "구체적으로, 카탈로그 모드는 실험 B(`human_20`)에서 문서 재현율이 **1.000**인데 "
            f"실험 A(`qa_sample`)에서는 질문당 문서를 평균 **{fmt(catalog_answer['avg_selected_doc_count'], 2)}개**밖에 "
            "고르지 못한다 — 20문항 중 19문항에서 문서를 하나도 선정하지 못했다. "
            "실험 C의 임계값 분석은 `human_*` 질문에서만 돌았고 거기서는 카탈로그 게이트가 정상으로 보인다. "
            "**따라서 낮은 답변률의 원인은 아직 규명되지 않았다.** "
            "다음 할 일은 `qa_sample` 질문셋으로 임계값 분석(실험 C)을 다시 돌려보는 것이다.",
            "",
            "---",
            "",
            f"## 실험 A — 답변 품질과 비용 (`qa_*` 질문 {n}개)",
            "",
            f"평가셋 `{EVAL_SET_BY_SOURCE['sample_qa']}` · depth `answer` (VLM 전체 호출)",
            "",
            md_table(ANSWER_HEADERS, answer_summary_rows(answer)),
            "",
            "- **답변률**: 모델이 기권(abstain)하지 않고 실제로 답한 질문 수.",
            "- **정확도 token_f1**: 기대 답변과 생성 답변의 토큰 겹침. 답한 질문만 대상이라 "
            "답변률이 낮으면 높게 나오기 쉽다 — 답변률과 반드시 같이 볼 것.",
            "- **VLM 페이지**: 질문당 VLM에 넣은 고유 페이지 수. 캐시 상태와 무관한 비용 지표다.",
            "",
        ]

    if docs:
        n = docs["eval_item_count"]
        lines += [
            "---",
            "",
            f"## 실험 B — 문서 선정 정확도 (`human_*` 질문 {n}개)",
            "",
            f"평가셋 `{EVAL_SET_BY_SOURCE['human']}` · depth `docs` (VLM 호출 0회)"
            " · **실험 A와 질문셋이 다르다**",
            "",
            md_table(DOCS_HEADERS, docs_summary_rows(docs)),
            "",
            "- **doc_recall@k**: 정답 문서가 선정된 문서 목록에 들어있는 비율.",
            "- **거절 정확도**: 문서에 답이 없는 무관 질문을 올바로 거절한 비율.",
            "",
        ]

    if thresholds:
        rows = []
        for mode in MODE_ORDER:
            entry = thresholds.get(mode)
            if not entry:
                continue
            stat = entry["stat"]
            too_high = (
                stat["relevant_min"] is not None and stat["relevant_min"] < stat["current_gate"]
            )
            rows.append(
                [
                    f"{MODE_LABEL[mode]} (`{mode}`)",
                    f"`{entry['gate_attr']}`",
                    fmt(stat["current_gate"]),
                    fmt(stat["relevant_min"]) + (" ⚠️" if too_high else ""),
                    fmt(stat["irrelevant_max"]),
                    fmt(stat.get("threshold_at_95pct_relevant_retention")),
                ]
            )
        lines += [
            "---",
            "",
            "## 실험 C — 문서 선정 컷오프 임계값 (`human_*` 질문)",
            "",
            "현재 게이트가 관련 문서까지 걸러내고 있는지 확인한 결과다. "
            "실험 B와 같은 질문셋이며, **실험 A의 낮은 답변률은 이 표로 설명되지 않는다.**",
            "",
            md_table(
                [
                    "모드",
                    "게이트 파라미터",
                    "현재 값",
                    "관련 문서 최저 점수",
                    "무관 문서 최고 점수",
                    "권장 임계값",
                ],
                rows,
            ),
            "",
            "**관련 문서 최저 점수**가 **현재 값**보다 낮으면(⚠️) 그 게이트는 정답 문서를 버리고 있다. "
            "관련 문서 최저 점수와 무관 문서 최고 점수 사이가 비어 있을수록 안전하게 임계값을 내릴 수 있다.",
            "",
        ]
        sources = sorted({entry["source"] for entry in thresholds.values()})
        if len(sources) > 1:
            lines += [
                "> 모드별로 유효한 최신 측정을 골랐다. 초기 실행은 `filename_only` 점수를 "
                "하나도 수집하지 못해(n_relevant=0) 해당 모드만 재실행한 결과를 쓴다.",
                "",
            ]

    lines += [
        "---",
        "",
        "## 더 자세히",
        "",
        "- 질문 하나하나의 3모드 답변 전문 비교: [01_질문별_답변비교.md](01_질문별_답변비교.md)",
        "- 엑셀에서 필터·정렬: [results.xlsx](results.xlsx)",
        "- 실험 설계와 배경: [../REPORT.md](../REPORT.md)",
        "",
        "이 문서는 `python -m tools.summarize_results` 로 원본 JSON에서 생성된다. 직접 고치지 말 것.",
        "",
        "생성에 쓰인 원본:",
        "",
    ]
    for depth, compare in sorted(compares.items()):
        rel = compare["_path"].relative_to(ROOT)
        lines.append(f"- `{rel.as_posix()}` (depth={depth})")
    for source in sorted({entry["source"] for entry in thresholds.values()}):
        lines.append(f"- `{source}`")
    lines.append("")

    (RESULTS / "00_SUMMARY.md").write_text("\n".join(lines), encoding="utf-8")


# ------------------------------------------------------- 01_질문별_답변비교


def write_qa_comparison(compare: dict, expected: dict[str, dict]) -> None:
    grouped = group_by_question(compare)
    lines = [
        "# 질문별 답변 비교 (3개 모드)",
        "",
        f"평가셋 `{EVAL_SET_BY_SOURCE['sample_qa']}` · 질문 {len(grouped)}개 · depth `answer`",
        "",
        "각 질문마다 기대 답변을 먼저 싣고, 세 모드가 실제로 무엇을 답했는지 이어서 보여준다.",
        "",
    ]

    for idx, (qid, by_mode) in enumerate(grouped.items(), start=1):
        sample = next(iter(by_mode.values()))
        exp = expected.get(qid, {})
        lines += [
            "---",
            "",
            f"## {idx}. {sample['question']}",
            "",
            f"`{qid}` · 유형 {sample.get('question_type', '-')} · 분류 {sample.get('category', '-')}",
            "",
        ]
        if exp.get("expected_answer"):
            lines += ["**기대 답변**", "", f"> {exp['expected_answer']}", ""]
        if exp.get("expected_documents"):
            docs = ", ".join(f"`{d}`" for d in exp["expected_documents"])
            lines += [f"**기대 문서**: {docs}", ""]

        rows = []
        for mode in MODE_ORDER:
            item = by_mode.get(mode)
            if not item:
                continue
            rows.append(
                [
                    MODE_LABEL[mode],
                    "기권" if item["abstained"] else "답변",
                    fmt(item["answer_token_f1"]),
                    fmt(item["expected_answer_keyword_recall"]),
                    fmt(item["elapsed_seconds_total"], 1),
                    str(item["summary_pages_required"] + item["chunk_pages_required"]),
                ]
            )
        lines += [
            md_table(
                ["모드", "결과", "token_f1", "키워드 재현율", "소요시간(초)", "VLM 페이지"],
                rows,
            ),
            "",
        ]

        for mode in MODE_ORDER:
            item = by_mode.get(mode)
            if not item:
                continue
            lines.append(f"### {MODE_LABEL[mode]}")
            lines.append("")
            selected = item["selected_documents"]
            lines.append(
                "선정 문서: " + (", ".join(f"`{d}`" for d in selected) if selected else "_없음_")
            )
            lines.append("")
            answer_text = (item["final_answer"] or "").strip()
            if item["abstained"] or not answer_text:
                lines += [f"> _기권_ — {answer_text or '답변 없음'}", ""]
            else:
                lines += ["> " + answer_text.replace("\n", "\n> "), ""]

    (RESULTS / "01_질문별_답변비교.md").write_text("\n".join(lines), encoding="utf-8")


# ------------------------------------------------------------------- xlsx

HEADER_FILL = PatternFill("solid", fgColor="DDE5F0")


def style_sheet(ws, widths: list[int], wrap_cols: set[int] = frozenset()) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for col in wrap_cols:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_xlsx(
    compares: dict[str, dict], thresholds: dict[str, dict], expected: dict[str, dict]
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    answer = compares.get("answer")
    if answer:
        ws = wb.create_sheet("요약")
        ws.append(ANSWER_HEADERS)
        for row in answer_summary_rows(answer):
            ws.append([c.replace("`", "") for c in row])
        style_sheet(ws, [22, 10, 18, 16, 18, 18, 18])

        ws = wb.create_sheet("질문별상세")
        ws.append(
            [
                "질문ID",
                "질문",
                "분류",
                "유형",
                "모드",
                "결과",
                "답변",
                "기대 답변",
                "token_f1",
                "키워드 재현율",
                "소요시간(초)",
                "VLM 페이지",
                "선정 문서 수",
                "선정 문서",
            ]
        )
        for qid, by_mode in group_by_question(answer).items():
            exp = expected.get(qid, {})
            for mode in MODE_ORDER:
                item = by_mode.get(mode)
                if not item:
                    continue
                ws.append(
                    [
                        qid,
                        item["question"],
                        item.get("category"),
                        item.get("question_type"),
                        MODE_LABEL[mode],
                        "기권" if item["abstained"] else "답변",
                        item["final_answer"],
                        exp.get("expected_answer"),
                        item["answer_token_f1"],
                        item["expected_answer_keyword_recall"],
                        round(item["elapsed_seconds_total"], 1),
                        item["summary_pages_required"] + item["chunk_pages_required"],
                        item["selected_doc_count"],
                        ", ".join(item["selected_documents"]),
                    ]
                )
        style_sheet(ws, [10, 40, 14, 16, 16, 8, 60, 60, 10, 14, 13, 12, 12, 40], {7, 8, 14})

    docs = compares.get("docs")
    if docs:
        ws = wb.create_sheet("문서선정_요약")
        ws.append(DOCS_HEADERS)
        for row in docs_summary_rows(docs):
            ws.append([c.replace("`", "") for c in row])
        style_sheet(ws, [22, 24, 12, 20, 18, 18])

        ws = wb.create_sheet("문서선정_상세")
        ws.append(
            [
                "질문ID",
                "질문",
                "모드",
                "정답 문서 적중",
                "MRR",
                "거절 정확",
                "소요시간(초)",
                "선정 문서 수",
                "선정 문서",
                "기대 문서",
            ]
        )
        for qid, by_mode in group_by_question(docs).items():
            exp = expected.get(qid, {})
            for mode in MODE_ORDER:
                item = by_mode.get(mode)
                if not item:
                    continue
                ws.append(
                    [
                        qid,
                        item["question"],
                        MODE_LABEL[mode],
                        item["doc_recall_hit"],
                        item["doc_mrr"],
                        item["rejection_correct"],
                        round(item["elapsed_seconds_total"], 1),
                        item["selected_doc_count"],
                        ", ".join(item["selected_documents"]),
                        ", ".join(exp.get("expected_documents") or []),
                    ]
                )
        style_sheet(ws, [10, 40, 16, 14, 10, 12, 13, 12, 40, 40], {9, 10})

    if thresholds:
        ws = wb.create_sheet("임계값")
        ws.append(
            [
                "모드",
                "게이트 파라미터",
                "현재 값",
                "관련 문서 수",
                "무관 문서 수",
                "관련 최저",
                "관련 최고",
                "무관 최저",
                "무관 최고",
                "권장 임계값",
                "출처 파일",
            ]
        )
        for mode in MODE_ORDER:
            entry = thresholds.get(mode)
            if not entry:
                continue
            stat = entry["stat"]
            ws.append(
                [
                    MODE_LABEL[mode],
                    entry["gate_attr"],
                    stat["current_gate"],
                    stat["n_relevant"],
                    stat["n_irrelevant"],
                    stat["relevant_min"],
                    stat["relevant_max"],
                    stat["irrelevant_min"],
                    stat["irrelevant_max"],
                    stat.get("threshold_at_95pct_relevant_retention"),
                    entry["source"],
                ]
            )
        style_sheet(ws, [18, 30, 12, 12, 12, 12, 12, 12, 12, 16, 44])

    wb.save(RESULTS / "results.xlsx")


def main() -> None:
    RESULTS.mkdir(exist_ok=True)
    compares = collect_compares()
    if not compares:
        raise SystemExit("outputs/comparisons/ 에 compare_*.json 이 없다.")

    thresholds = pick_thresholds()
    expected = load_expected_answers()

    write_summary(compares, thresholds)
    if "answer" in compares:
        write_qa_comparison(compares["answer"], expected)
    write_xlsx(compares, thresholds, expected)

    print(f"생성 완료 → {RESULTS}")
    for path in sorted(RESULTS.iterdir()):
        print(f"  {path.name}")


if __name__ == "__main__":
    main()
